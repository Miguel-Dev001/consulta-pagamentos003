# Bibliotecas
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from pathlib import Path
import os

# Entrar na planilha e extrair dados
planilha_path = os.getcwd() + '/dados_clientes.xlsx'
print(planilha_path)
planilha_clientes = openpyxl.load_workbook(planilha_path)
pagina_clientes = planilha_clientes['Sheet1']

# Entrar no site
driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/')
sleep(5)

for linha in pagina_clientes.iter_rows(min_row = 2,values_only=True):
    nome, valor, CPF, vencimento = linha 
    
# Campo 'Digite o CPF'
    campo_pesquisa = driver.find_element(By.XPATH,"//input[@id='cpfInput']")
    sleep(3)
    campo_pesquisa.clear()
    campo_pesquisa.send_keys(CPF)
    sleep(3)

# Clicar em consultar
    botao_consultar = driver.find_element(By.XPATH,"//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(1)
    botao_consultar.click()
    sleep(4)
    
# Status
    status = driver.find_element(By.XPATH,"//span[@id='statusLabel']")
    if status.text == 'em dia':
# Se estiver "em dia",pegar a data do pagamento e o método de pagamento
        data_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentDate']")
        metodo_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentMethod']")
    
        data_pagamento_limpo = data_pagamento.text.split()[3]
        metodo_pagamento_limpo = metodo_pagamento.text.split()[3]
    
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']

        pagina_fechamento.append([nome, valor, CPF, vencimento,'em dia', data_pagamento_limpo, metodo_pagamento_limpo]) 
    
        planilha_fechamento.save('planilha fechamento.xlsx') 
    else:
# Caso estiver "atrasado" colocar status como pendente
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']
        
        pagina_fechamento.append([nome, valor, CPF, vencimento,'pendente'])
        planilha_fechamento.save('planilha fechamento.xlsx')